# formatter.py — 招聘猫 格式化导出模块
import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List

from models import Notice, ScanResult

logger = logging.getLogger(__name__)


def export_to_excel(results: List[ScanResult], filepath: str, only_new: bool = False):
    """导出到 Excel（.xlsx）"""
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
    except ImportError as e:
        raise ImportError("请先安装 openpyxl：pip install openpyxl") from e

    wb = openpyxl.Workbook()

    # ── 汇总 Sheet ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "汇总"
    _write_summary_sheet(ws_summary, results)

    # ── 新通知 Sheet ────────────────────────────────────────────────────
    ws_new = wb.create_sheet("新通知")
    notices_to_write: List[tuple] = []
    for r in results:
        src = r.new_notices if only_new else r.all_notices
        for n in src:
            status = "新增" if n in r.new_notices else ("消失" if n in r.gone_notices else "未变")
            notices_to_write.append((r.website, status, n))

    _write_notice_sheet(ws_new, notices_to_write)

    # ── 每站独立 Sheet ──────────────────────────────────────────────────
    for r in results:
        safe_name = r.website.replace("https://", "").replace("http://", "")[:28]
        ws = wb.create_sheet(safe_name)
        rows = [(r.website, "新增" if n in r.new_notices else "未变", n) for n in r.all_notices]
        rows += [(r.website, "已消失", n) for n in r.gone_notices]
        _write_notice_sheet(ws, rows)

    wb.save(filepath)
    logger.info(f"Excel 已保存到 {filepath}")


def _write_summary_sheet(ws, results: List[ScanResult]):
    from openpyxl.styles import Font, PatternFill, Alignment
    headers = ["网站", "扫描时间", "总通知数", "新增", "消失", "未变", "状态", "错误信息"]
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        status = "🆕 有新通知" if r.has_new else ("❌ 出错" if r.error else "✅ 无变化")
        ws.cell(row=row_idx, column=1, value=r.website)
        ws.cell(row=row_idx, column=2, value=r.scan_time)
        ws.cell(row=row_idx, column=3, value=len(r.all_notices))
        ws.cell(row=row_idx, column=4, value=len(r.new_notices))
        ws.cell(row=row_idx, column=5, value=len(r.gone_notices))
        ws.cell(row=row_idx, column=6, value=len(r.unchanged_notices))
        ws.cell(row=row_idx, column=7, value=status)
        ws.cell(row=row_idx, column=8, value=r.error or "")

        if r.has_new:
            ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor="FFEB9C")
            ws.cell(row=row_idx, column=7).font = Font(color="9C6500", bold=True)

    col_widths = [45, 20, 10, 8, 8, 8, 15, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def _write_notice_sheet(ws, rows: List[tuple]):
    """rows: [(website, status, Notice)]"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    headers = ["来源网站", "状态", "标题", "链接", "日期", "摘要", "抓取时间"]
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    new_fill = PatternFill("solid", fgColor="E2EFDA")
    gone_fill = PatternFill("solid", fgColor="FCE4D6")

    for row_idx, (website, status, n) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=website)
        status_cell = ws.cell(row=row_idx, column=2, value=status)
        ws.cell(row=row_idx, column=3, value=n.title)
        link_cell = ws.cell(row=row_idx, column=4, value=n.link)
        link_cell.hyperlink = n.link
        link_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_idx, column=5, value=n.date)
        ws.cell(row=row_idx, column=6, value=n.snippet)
        ws.cell(row=row_idx, column=7, value=n.scraped_at)

        if status == "新增":
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = new_fill
        elif status == "已消失":
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = gone_fill

    col_widths = [40, 8, 60, 60, 14, 50, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    ws.auto_filter.ref = ws.dimensions


def export_to_csv(results: List[ScanResult], filepath: str):
    """导出到 CSV"""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["来源网站", "状态", "标题", "链接", "日期", "摘要", "抓取时间"])
        for r in results:
            new_hashes = {n.content_hash for n in r.new_notices}
            gone_hashes = {n.content_hash for n in r.gone_notices}
            for n in r.all_notices:
                status = "新增" if n.content_hash in new_hashes else "未变"
                writer.writerow([r.website, status, n.title, n.link, n.date, n.snippet, n.scraped_at])
            for n in r.gone_notices:
                writer.writerow([r.website, "已消失", n.title, n.link, n.date, n.snippet, n.scraped_at])
    logger.info(f"CSV 已保存到 {filepath}")
